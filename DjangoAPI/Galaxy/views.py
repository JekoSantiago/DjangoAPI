from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.http.response import JsonResponse,HttpResponse,StreamingHttpResponse
from django.db import connection, connections
from django.template import loader
from DjangoAPI.helper import query_db
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


# def query_db(query, args=(), one=False):
#     cur = connections['galaxy'].cursor()
#     cur.execute(query, args)
#     r = [dict((cur.description[i][0], value) 
#         for i, value in enumerate(row)) for row in cur.fetchall()]
#     cur.connection.close()
#     return (r[0] if r else None) if one else r

@csrf_exempt
def operatorGet(request):
    data = query_db('galaxy','SELECT * FROM "Operator" ORDER BY "Operator_ID" ASC')
    return JsonResponse(data,safe=False)

@csrf_exempt
def operatorInsert(request):
    operatorName = request.POST.get('operatorName')
    userID = request.POST.get('userID')
    data = query_db('galaxy','INSERT INTO "Operator"("OperatorName","InsertBy","LastUpdateBy","LastUpdate") VALUES(\'' + str(operatorName) +'\','+userID+','+userID+', NOW() ) RETURNING "Operator_ID" AS RETURN, \'Succefully Added Operator\' AS MESSAGE')
    return JsonResponse(data,safe=False)

@csrf_exempt
def operatorUpdate(request):
    operatorName = request.POST.get('operatorName')
    userID = request.POST.get('userID')
    status = request.POST.get('isActive')
    operatorID = request.POST.get('operatorID')
    data = query_db('galaxy','UPDATE "Operator" SET "OperatorName" =\''+ str(operatorName)+'\', "LastUpdateBy" = ' +userID+', "LastUpdate" = NOW(), "isActive" = ' + status +' WHERE "Operator_ID" = ' + operatorID +' RETURNING "Operator_ID" AS RETURN, \'Succefully Updated Operator\' AS MESSAGE')
    return JsonResponse(data,safe=False)



def export_page(request):    
    collection = query_db('galaxy','SELECT * FROM "Operator" ORDER BY "Operator_ID" ASC')
    output = BytesIO()
    df = pd.DataFrame(collection)
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    output.seek(0)
    filename = "ExportXLS"
    response = StreamingHttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    return response

def viewxls(request):
    data = query_db('galaxy','SELECT * FROM "Operator" ORDER BY "Operator_ID" ASC')
    wb = Workbook()
    ws = wb.active

    #TITLE
    ws.merge_cells('A1:F1')
    top_left_cell = ws['A1']
    top_left_cell.value = "Operators"
    thin = Side(border_style="thin", color="000000")
    top_left_cell.font  = Font(b=True, color="000000", size=16)
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    #TABLE HEADERS
    ws['A3'] = "OPERATOR ID"
    ws['B3'] = "OPERATOR NAME"
    ws['C3'] = "STATUS"
    ws['A3'].font = Font(b=True)
    ws['B3'].font = Font(b=True)
    ws['C3'].font = Font(b=True)

    #CONTENT
    lastRow=4
    for item in data:
        for key,value in item.items():
            if key == "Operator_ID":
                ws["A"+str(lastRow)] = value
            if key == "OperatorName":
                ws["B"+str(lastRow)] = value    
            if key == "isActive":
                ws["C"+str(lastRow)] = value
        lastRow +=1

    #EXPORT DOWNLOAD
    filename = "ExportXLS"
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    return response        
